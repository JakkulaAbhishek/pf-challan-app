import streamlit as st
import pdfplumber
import re
import pandas as pd
from io import BytesIO
from openpyxl.styles import Font, Alignment, PatternFill
from datetime import datetime
import plotly.express as px
from fpdf import FPDF

# ---------------- PAGE CONFIG ----------------
st.set_page_config(page_title="PF AI Auditor | Abhishek Jakkula", layout="wide")

# ---------------- ULTRA STYLISH UI THEME ----------------
st.markdown("""
<style>
    @import url('https://fonts.googleapis.com/css2?family=Poppins:wght@300;400;600;800&display=swap');
    html, body, [class*="css"] { font-family: 'Poppins', sans-serif; }
    .stApp { background: radial-gradient(circle at top, #0f172a, #020617); color: #f8fafc; }
    .header-card {
        background: linear-gradient(135deg, rgba(30, 41, 59, 0.7), rgba(15, 23, 42, 0.8));
        padding: 40px; border-radius: 28px; border: 1px solid rgba(255, 255, 255, 0.1);
        backdrop-filter: blur(12px); text-align: center; margin-bottom: 40px;
    }
    .main-title {
        background: linear-gradient(90deg, #38bdf8, #818cf8, #34d399);
        -webkit-background-clip: text; -webkit-text-fill-color: transparent;
        font-weight: 800; font-size: 3.5rem; line-height: 1.1;
    }
    .branding { color: #94a3b8; font-size: 1.1rem; letter-spacing: 2px; margin-top: 15px; font-weight: 600; }
    .stButton>button {
        background: linear-gradient(90deg, #2563eb, #0ea5e9);
        color: white !important; border: none; border-radius: 14px;
        font-weight: 800; height: 60px; width: 100%; transition: 0.4s;
    }
</style>
""", unsafe_allow_html=True)

# ---------------- BRANDED HEADER ----------------
st.markdown("""
<div class="header-card">
    <div class="main-title">PF CHALLAN AI AUDITOR</div>
    <div class="branding">AUDIT TRAIL & COMPLIANCE BY <b>ABHISHEK JAKKULA</b></div>
    <div style="color: #facc15; font-style: italic; margin-top: 15px;">🌸 Lord Krishna Blessings: कर्मण्येवाधिकारस्ते मा फलेषु कदाचन 🌸</div>
</div>
""", unsafe_allow_html=True)

# ---------------- HELPERS ----------------
def safe_extract(pattern, text):
    m = re.search(pattern, text, re.I | re.S)
    return m.group(1).strip() if m else "0"

def calculate_due_date(wage_month_str):
    try:
        parts = wage_month_str.split()
        month_dt = datetime.strptime(parts[0], "%B")
        year = int(parts[1])
        next_m = month_dt.month % 12 + 1
        next_y = year + (1 if month_dt.month == 12 else 0)
        return datetime(next_y, next_m, 15)
    except: return None

# ---------------- PDF GENERATOR (AUDIT TRAIL) ----------------
def generate_pdf_summary(df, total_pf, disallowance, late_count):
    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("Arial", 'B', 16)
    pdf.cell(200, 10, "PF COMPLIANCE AUDIT CERTIFICATE", ln=True, align='C')
    pdf.set_font("Arial", '', 10)
    pdf.cell(200, 10, f"Generated on: {datetime.now().strftime('%d-%m-%Y %H:%M')}", ln=True, align='C')
    pdf.ln(10)
    
    # Summary Table
    pdf.set_font("Arial", 'B', 12)
    pdf.cell(100, 10, "Executive Summary", ln=True)
    pdf.set_font("Arial", '', 11)
    pdf.cell(100, 8, f"Total Challans Processed: {len(df)}")
    pdf.ln(8)
    pdf.cell(100, 8, f"Total PF Amount Paid: INR {total_pf:,.2f}")
    pdf.ln(8)
    pdf.cell(100, 8, f"Total Late Filings Detected: {late_count}")
    pdf.ln(8)
    pdf.set_text_color(255, 0, 0)
    pdf.cell(100, 8, f"Statutory Disallowance (U/s 36(1)(va)): INR {disallowance:,.2f}")
    pdf.set_text_color(0, 0, 0)
    pdf.ln(15)

    # Detailed Table Header
    pdf.set_font("Arial", 'B', 10)
    pdf.set_fill_color(240, 240, 240)
    pdf.cell(40, 8, "Wage Month", 1, 0, 'C', True)
    pdf.cell(40, 8, "Due Date", 1, 0, 'C', True)
    pdf.cell(40, 8, "Paid Date", 1, 0, 'C', True)
    pdf.cell(35, 8, "Late Days", 1, 0, 'C', True)
    pdf.cell(35, 8, "Disallowance", 1, 1, 'C', True)

    # Table Rows
    pdf.set_font("Arial", '', 9)
    for index, row in df.iterrows():
        pdf.cell(40, 7, str(row['Wage Month']), 1)
        pdf.cell(40, 7, str(row['Due Date']), 1)
        pdf.cell(40, 7, str(row['Generated Date']), 1)
        pdf.cell(35, 7, str(row['Late Days']), 1, 0, 'C')
        pdf.cell(35, 7, f"{row['PF Disallowance (₹)']:,.2f}", 1, 1, 'R')

    pdf.ln(20)
    pdf.set_font("Arial", 'I', 10)
    pdf.cell(200, 10, "This is an AI-generated audit trail for internal compliance verification.", ln=True, align='C')
    pdf.cell(200, 10, "Digitally Verified by Abhishek Jakkula Audit Engine", ln=True, align='C')
    
    return pdf.output(dest='S').encode('latin-1')

# ---------------- MAIN APP ----------------
files = st.file_uploader("📂 Upload PF Challan PDFs", type="pdf", accept_multiple_files=True)
run_audit = st.button("🚀 INITIATE PF AI AUDIT")

if files and run_audit:
    all_rows = []
    for f in files:
        with pdfplumber.open(f) as pdf:
            text = "\n".join([p.extract_text() for p in pdf.pages if p.extract_text()])
            blocks = re.split(r"(Dues for the wage month of\s*[A-Za-z]+\s*[0-9]{4})", text, flags=re.I)
            for i in range(1, len(blocks), 2):
                content = blocks[i] + blocks[i+1]
                m_match = re.search(r"wage month of\s*([A-Za-z]+)\s*([0-9]{4})", content, re.I)
                wage_month = f"{m_match.group(1).title()} {m_match.group(2)}" if m_match else "Unknown"
                sys_date_str = safe_extract(r"system generated challan on\s*.*?(\d{2}-[A-Z]{3}-\d{4})", content).upper()
                due_dt = calculate_due_date(wage_month)
                late_days, disallowance, status = 0, 0.0, "On Time ✅"
                emp_share = float(safe_extract(r"Employee'?s Share Of.*?([0-9,]{2,})", content).replace(",", ""))
                if due_dt and sys_date_str != "0":
                    try:
                        gen_dt = datetime.strptime(sys_date_str, "%d-%b-%Y")
                        if gen_dt > due_dt:
                            status, late_days, disallowance = "Late Payment ⚠️", (gen_dt - due_dt).days, emp_share
                    except: pass
                all_rows.append({
                    "Wage Month": wage_month, "Due Date": due_dt.strftime("%d-%b-%Y") if due_dt else "Unknown",
                    "Generated Date": sys_date_str, "Payment Status": status, "Late Days": late_days,
                    "Admin Charges (₹)": float(safe_extract(r"Administration Charges.*?([0-9,]{2,})", content).replace(",", "")),
                    "Employer Share (₹)": float(safe_extract(r"Employer'?s Share Of.*?([0-9,]{2,})", content).replace(",", "")),
                    "Employee Share (₹)": emp_share, "PF Disallowance (₹)": disallowance,
                    "Grand Total (₹)": float(safe_extract(r"Grand Total.*?([0-9,]{2,})", content).replace(",", "")), "Source File": f.name
                })

    if all_rows:
        df = pd.DataFrame(all_rows)
        total_pf = df['Grand Total (₹)'].sum()
        total_dis = df['PF Disallowance (₹)'].sum()
        late_count = len(df[df['Late Days'] > 0])

        st.markdown("### 📊 AUDIT COMMAND DASHBOARD")
        m1, m2, m3 = st.columns(3)
        m1.metric("Total PF Audited", f"₹{total_pf:,.2f}")
        m2.metric("PF Disallowance", f"₹{total_dis:,.2f}", delta=f"Risk: ₹{total_dis*0.3:,.0f}")
        m3.metric("Late Filings", late_count)

        st.markdown("---")
        c1, c2 = st.columns(2)
        with c1:
            st.download_button("🚀 DOWNLOAD EXCEL AUDIT", data=df.to_csv().encode('utf-8'), file_name="PF_Audit.csv")
        with c2:
            pdf_data = generate_pdf_summary(df, total_pf, total_dis, late_count)
            st.download_button("📜 DOWNLOAD PDF AUDIT TRAIL", data=pdf_data, file_name="PF_Audit_Trail.pdf", mime="application/pdf")

        st.dataframe(df.style.format({"Grand Total (₹)": "{:,.2f}", "PF Disallowance (₹)": "{:,.2f}"}), use_container_width=True)

st.caption("⚙️ Auditor Pro developed by Abhishek Jakkula")
